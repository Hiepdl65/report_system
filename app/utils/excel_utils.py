from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class ExcelFormatter:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def format_worksheet(self, worksheet, workbook, dataframe: pd.DataFrame):
        """Apply formatting to Excel worksheet"""
        # Format headers
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all cells with data
        for row in worksheet.iter_rows(min_row=1, max_row=len(dataframe) + 1):
            for cell in row:
                cell.border = self.border
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions